"""Financial data parser for Excel files."""

import io
from typing import Dict, List, Tuple, Optional
import pandas as pd
import openpyxl
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

# Use absolute imports
from packages.core.models.financials import FinancialSeries, FinancialsData


def parse_financials(xlsx_bytes: bytes) -> FinancialsData:
    """
    Parse financial data from Excel file.
    
    Args:
        xlsx_bytes: Raw Excel file bytes
        
    Returns:
        FinancialsData with extracted series
    """
    workbook = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    
    series_list = []
    
    # Try to find financial data in common sheet names
    financial_sheets = ["Financials", "Data", "Revenue", "Financial Data", "Sheet1"]
    
    for sheet_name in financial_sheets:
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            series = _extract_from_sheet(sheet, sheet_name)
            series_list.extend(series)
    
    # If no data found, try all sheets
    if not series_list:
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            series = _extract_from_sheet(sheet, sheet_name)
            series_list.extend(series)
    
    # Try to extract from named ranges
    named_range_series = _extract_from_named_ranges(workbook)
    series_list.extend(named_range_series)
    
    return FinancialsData(series=series_list)


def _extract_from_sheet(sheet: Worksheet, sheet_name: str) -> List[FinancialSeries]:
    """Extract financial series from a worksheet."""
    series_list = []
    
    # Look for common financial metrics in column headers
    for col in range(1, sheet.max_column + 1):
        header_cell = sheet.cell(row=1, column=col)
        header_text = str(header_cell.value).strip().lower()
        
        if _is_financial_metric(header_text):
            series_data = _extract_column_data(sheet, col, header_text)
            if series_data:
                series = FinancialSeries(
                    name=_normalize_metric_name(header_text),
                    data=series_data,
                    sheet_name=sheet_name
                )
                series_list.append(series)
    
    # Look for year-based data (common pattern: years in first row, metrics in first column)
    year_based_series = _extract_year_based_data(sheet, sheet_name)
    series_list.extend(year_based_series)
    
    return series_list


def _is_financial_metric(text: str) -> bool:
    """Check if text represents a financial metric."""
    financial_keywords = [
        "revenue", "sales", "ebitda", "ebit", "net income", "profit",
        "cash flow", "assets", "liabilities", "equity", "debt",
        "capex", "opex", "gross margin", "operating margin"
    ]
    
    return any(keyword in text for keyword in financial_keywords)


def _normalize_metric_name(text: str) -> str:
    """Normalize metric name to standard format."""
    # Map common variations to standard names
    name_mapping = {
        "revenue": "Revenue",
        "sales": "Revenue",
        "ebitda": "EBITDA",
        "ebit": "EBIT",
        "net income": "Net Income",
        "profit": "Net Income",
        "cash flow": "Cash Flow",
        "assets": "Total Assets",
        "liabilities": "Total Liabilities",
        "equity": "Total Equity",
        "debt": "Total Debt",
        "capex": "Capital Expenditure",
        "opex": "Operating Expenses",
        "gross margin": "Gross Margin",
        "operating margin": "Operating Margin"
    }
    
    for key, value in name_mapping.items():
        if key in text.lower():
            return value
    
    # Capitalize first letter of each word
    return " ".join(word.capitalize() for word in text.split())


def _extract_column_data(sheet: Worksheet, col: int, metric_name: str) -> List[Tuple[int, float]]:
    """Extract data from a column, assuming years in first column."""
    data = []
    
    # Look for years in first column
    for row in range(2, sheet.max_row + 1):
        year_cell = sheet.cell(row=row, column=1)
        value_cell = sheet.cell(row=row, column=col)
        
        year = _extract_year(year_cell.value)
        value = _extract_numeric_value(value_cell.value)
        
        if year and value is not None:
            data.append((year, value))
    
    return data


def _extract_year(value) -> Optional[int]:
    """Extract year from cell value."""
    if value is None:
        return None
    
    try:
        # Try to convert to int
        year = int(value)
        if 1900 <= year <= 2100:
            return year
    except (ValueError, TypeError):
        pass
    
    # Try to extract year from date
    try:
        if hasattr(value, 'year'):
            return value.year
    except AttributeError:
        pass
    
    # Try to extract year from string
    try:
        if isinstance(value, str):
            import re
            year_match = re.search(r'\b(19|20)\d{2}\b', value)
            if year_match:
                return int(year_match.group())
    except (ValueError, TypeError):
        pass
    
    return None


def _extract_numeric_value(value) -> Optional[float]:
    """Extract numeric value from cell."""
    if value is None:
        return None
    
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def _extract_year_based_data(sheet: Worksheet, sheet_name: str) -> List[FinancialSeries]:
    """Extract data assuming years are in first row and metrics in first column."""
    series_list = []
    
    # Check if first row contains years
    years = []
    for col in range(2, sheet.max_column + 1):
        year = _extract_year(sheet.cell(row=1, column=col).value)
        if year:
            years.append(year)
    
    if not years:
        return []
    
    # Extract metrics from first column
    for row in range(2, sheet.max_row + 1):
        metric_cell = sheet.cell(row=row, column=1)
        metric_name = str(metric_cell.value).strip()
        
        if _is_financial_metric(metric_name):
            data = []
            for col, year in enumerate(years, start=2):
                value = _extract_numeric_value(sheet.cell(row=row, column=col).value)
                if value is not None:
                    data.append((year, value))
            
            if data:
                series = FinancialSeries(
                    name=_normalize_metric_name(metric_name),
                    data=data,
                    sheet_name=sheet_name
                )
                series_list.append(series)
    
    return series_list


def _extract_from_named_ranges(workbook: Workbook) -> List[FinancialSeries]:
    """Extract data from Excel named ranges."""
    series_list = []
    
    for name, named_range in workbook.defined_names.items():
        try:
            # Try to extract data from named range
            if hasattr(named_range, 'destinations'):
                for sheet_name, cell_range in named_range.destinations:
                    if sheet_name in workbook.sheetnames:
                        sheet = workbook[sheet_name]
                        data = _extract_from_named_range(sheet, cell_range, name)
                        if data:
                            series = FinancialSeries(
                                name=_normalize_metric_name(name),
                                data=data,
                                sheet_name=sheet_name,
                                range_name=name
                            )
                            series_list.append(series)
        except Exception:
            # Skip named ranges that can't be processed
            continue
    
    return series_list


def _extract_from_named_range(sheet: Worksheet, cell_range: str, range_name: str) -> List[Tuple[int, float]]:
    """Extract data from a named range."""
    try:
        # Parse cell range (e.g., "A1:B10")
        from openpyxl.utils import range_boundaries
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
        
        data = []
        
        # Assume first column contains years, second column contains values
        for row in range(min_row, max_row + 1):
            year = _extract_year(sheet.cell(row=row, column=min_col).value)
            value = _extract_numeric_value(sheet.cell(row=row, column=min_col + 1).value)
            
            if year and value is not None:
                data.append((year, value))
        
        return data
    except Exception:
        return []
